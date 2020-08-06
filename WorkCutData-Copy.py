#-*- codeing = utf-8 -*-
#@Time :2020/8/4 21:12
#@Author ：谈陆军
#@File ：WorkCutData-Copy.py
#@Software :PyCharm

from openpyxl import load_workbook
from openpyxl import Workbook
import xlrd
import requests
import sqlite3



url = "http://openapi.uml-tech.com"
Carfarm = "/farm/getHistoryWorkInfo"
URL = url + Carfarm


dbpath = "作业段是否上传.xlsx"
datalist04 = []


headers = {
    'Content-Type': 'application/json'
}

workbok = xlrd.open_workbook("车辆列表.xls")
sheet1 = workbok.sheet_by_index(0)


#共几行，共几列
nrows = sheet1.nrows
ncols = sheet1.ncols

row_values = sheet1.row_values(rowx=0)
print("共%d行%d列"%(nrows,ncols))



col = ['GPSDateTime','brand','jobArea','wktPoly','jobStartLat','vehicleTypeCode','jobEquipment','city','jobStartTime','jobWidth','currentBlockNumber','jobType','jobStartLon','did','jobEndTime','lk']
datalist = []
datalist02 = []

for i in range(0, nrows):
    record = sheet1.row_values(rowx=i)
    tows = record[1]


    payload = {"did": tows, "gpsEndTime": "20191231235959", "pageSize": 900, "pageType": 0, "reversed": "false",
               "startRowkey": "", "gpsStartTime": "20190101000000", "token": "92d4e3fef54e432a01127200365d039f"}
    payload = str(payload)



    response = requests.request("POST", URL, headers=headers, data=payload)
    DataBase = response.json()
    if DataBase['result'] != []:
        database = DataBase['result']
        datalist03 = []
        for j in range(len(database)):
            database02 = database[j]['work']

            datalist01 = []
            for deta01 in database02.keys():
                #print(deta01)
                if deta01 in col:
                    if database02[deta01]==[]:
                        datalist01.append(None)
                        continue
                    datalist01.append(database02[deta01])
            datalist03.append(datalist01)
        datalist.append(datalist03)


for m in range(0,len(datalist)):
    for k in range(0, len(datalist[m])):
        datalist04.append(datalist[m][k])

print("...加载完毕...")


save_path = r"D:\PycharmProjects\learnpython\untitled"+r"\作业段数据.xlsx"
wb = Workbook()
wb.create_sheet("sheet01")

ws = wb.active


for i in range(1,len(datalist04)):
    data = datalist04[i-1]
    if len(data)==16:
        for j in range(1,17):
            ws.cell(i, j,str(data[j-1]))

wb.save(save_path)
print("。。。存储完毕。。。")




