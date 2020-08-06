#-*- codeing = utf-8 -*-
#@Time :2020/7/18 16:50
#@Author ：谈陆军
#@File ：CarInfo.py
#@Software :PyCharm

import xlrd
import requests
import sqlite3
from openpyxl import load_workbook
from openpyxl import Workbook


url = "http://openapi.uml-tech.com"
Carfarm = "/farm/getVehicleInfo"
URL = url + Carfarm

dbpath = "车辆信息.xlsx"

datalist = []

headers = {
    'Content-Type': 'application/json'
}

workbok = xlrd.open_workbook("车辆列表.xls")
sheet1 = workbok.sheet_by_index(0)


#共几行，共几列
nrows = sheet1.nrows
ncols = sheet1.ncols

row_values = sheet1.row_values(rowx=0)
print("共%d行%d列"%(nrows,ncols))

for i in range(0,nrows):
    record = sheet1.row_values(rowx=i)
    hows = record[0]
    tows = record[1]
    datalit01 = []

    payload = {"dids": tows, "token": "92d4e3fef54e432a01127200365d039f", "vins": hows}

    payload = str(payload)

    response = requests.request("POST", URL, headers=headers, data=payload)
    DataBase = response.json()
    DataBase01 = DataBase["result"][0]
    for deta in DataBase01.values():
        if len(DataBase01.values())==6:
            if deta == "1":
                datalit01.append("1")
                datalit01.append("wu")
            else:
                datalit01.append(deta)
        else:
            datalit01.append(deta)
    datalist.append(datalit01)
print("...加载完毕...")



save_path = r"D:\PycharmProjects\learnpython\untitled"+r"\车辆信息.xlsx"
wb = Workbook()
wb.create_sheet("sheet01")

ws = wb.active

for i in range(1,len(datalist)):
    data = datalist[i-1]
    for j in range(1,8):
        ws.cell(i, j,str(data[j-1]))

wb.save(save_path)
print("。。。存储完毕。。。")





#
# def saveData2DB(datalist,dbpath):
#     init_db(dbpath)
#     conn = sqlite3.connect(dbpath)
#     cur = conn.cursor()
#
#     for data in datalist:
#         if len(data) < 7:
#             continue
#         for index in range(len(data)):
#             if index==2:
#                 continue
#             data[index] = "'"+data[index]+"'"
#         sql = '''
#                 insert into carInfo(
#                 did,groupName,groupType,lpn,vehicleModelCode,vehicleModelName,vin)
#                 values(%s)'''%",".join(data)
#         cur.execute(sql)
#         conn.commit()
#     cur.close()
#     conn.close()
#     print("存储完毕")
#
#
#
# def init_db(dbpath):
#     sql = '''
#         create table carInfo
#         (
#         did text primary key,
#         groupName text,
#         groupType integer,
#         lpn text,
#         vehicleModelCode text,
#         vehicleModelName text,
#         vin text)
#     '''
#     conn = sqlite3.connect(dbpath)
#     cursor = conn.cursor()
#     cursor.execute(sql)
#     conn.commit()
#     conn.close()
#
# saveData2DB(datalist,dbpath)