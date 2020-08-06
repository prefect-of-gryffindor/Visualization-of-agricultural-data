#-*- codeing = utf-8 -*-
#@Time :2020/7/28 20:13
#@Author ：谈陆军
#@File ：test-test01.py
#@Software :PyCharm


from openpyxl import load_workbook
from openpyxl import Workbook
import requests
import sqlite3




url = "http://openapi.uml-tech.com"
Carfarm = "/farm/getCoodinateFlag"
URL = url+Carfarm



datalist04 = []
datalist03 = []


headers = {
  'Content-Type': 'application/json'
}



for i in range(2019,2020):
    for j in range(1,13):
        if j<10:
            Time = str(i)+'-'+'0'+str(j)+'-01'
        else:
            Time = str(i)+'-'+str(j)+'-01'


        payload = {"date": Time,"token": "92d4e3fef54e432a01127200365d039f"}
        payload = str(payload)

        response = requests.request("POST", URL, headers=headers, data = payload)
        DataBase = response.json()

        DataBase01 = DataBase["result"][0]

        datalist02 = []
        for deta01 in DataBase01.keys():
            # sheet.write(count, 0, deta01)
            # y = 0
            datalist = []
            datalist01 = []
            for deta02 in list(DataBase01[deta01]):
                datalist.append(deta01)
                datalist.append(deta02)


                datalist01.append(datalist)
                datalist = []

            datalist02.append(datalist01)
        datalist04.append(datalist02)
    #print(datalist04)

for m in range(0,12):
    for k in range(0, len(datalist04[m])):
        for z in range(0,len(datalist04[m][k])):
            datalist03.append(datalist04[m][k][z])
        # print(len(datalist04[m][k]))

# print(len(datalist03))


save_path = r"D:\PycharmProjects\learnpython\untitled"+r"\上传车次.xlsx"
wb = Workbook()
wb.create_sheet("car_uplod")

ws = wb.active
ws["A1"] = "时间"
ws["B1"] = "设备号"

for i in range(1,len(datalist03)):
    data = datalist03[i-1]
    for j in range(1,3):
        ws.cell(i, 1,str(data[0]))
        ws.cell(i, 2,str(data[1]))


wb.save(save_path)
print("。。。存储完毕。。。")





























