import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import gc

gc.collect()
url = "http://openapi.uml-tech.com/farm/getLocation"
header = {'Content-Type': 'application/json'}
#------------获取50辆农机的did号----------------
excel_did = load_workbook('50_did.xlsx')
table_did = excel_did['Sheet1']
did_range = table_did['c2':'c51']
dids = ''
counts = 0
for i in did_range:
    counts+=1
    if counts!=len(did_range):
        dids += str(i[0].value) + ','
    else:
        dids += str(i[0].value)
# #---------------获取要加入数据库的字段编号、名称--------------
# excel_filed = load_workbook('filed.xlsx')
# table_filed = excel_filed['filed']
# filed_range = table_filed['A2':'C32']
# #-----------------创建excel表格并将字段名加入表头--------------
# wb = Workbook()
# ws = wb.active
# title = []
# for i in range(len(filed_range)):
#     title.append(filed_range[i][2].value)
# ws.append(title)
# wb.save('data(0-50).xlsx')
#--------------------从接口获取数据---------------------------
count = 0
# date = {
#     "1217",
#     "1218",
#     "1231"
# }
#按车辆循环获取
body = "{    \"did\": \"YMJ4121905100062\"," \
          "    \"token\": \"92d4e3fef54e432a01127200365d039f\"," \
          "    \"vin\": \"\"\r\n}"
response = requests.post(url=url , headers=header , data=body)
count +=1
print("这是向农机can数据接口的第：{0}次发送".format(count))
if(response.status_code==200):
    print("请求成功")
else:
    print("请求失败")
print(response.text.encode('utf8'))


